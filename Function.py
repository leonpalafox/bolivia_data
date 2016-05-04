# Funcion.py
import openpyxl
import numpy as np
import matplotlib.pyplot as plt

from sklearn import manifold
from sklearn import preprocessing
from sklearn.cluster import MeanShift

################	   Excel	     ################
wb = openpyxl.load_workbook('/Test/Data Base Bolivia.xlsx')
sheet = wb.get_sheet_by_name('Pooled data Bolivia')
sheet0 = wb.get_sheet_by_name('Item_Explanation')
#####################################################
def tabla (Min, Max, Start, Stop, Var):
    
    a = 0
    n = []
    d = []
    
    c = np.zeros((1, Stop-Start))
    c = np.delete(c, 0, 0)
    
    for i in range (Min, Max, 1):
        a = 0
        for j in range (Start, Stop, 1):
            if sheet.cell(row=i, column=j).value is None or sheet.cell(row=i, column=Var).value is None:
                a = 1
            else:
                d.append(sheet.cell(row=i, column=j).value)  
                
        if a == 0:
            c = np.vstack((c,d))
            n.append(i)
            
        d = []
        
    return (c, n)

def armar (Vector, Cluster):
    Cluster = np.vstack((Cluster, Vector))
    return Cluster
    
def desarmar (Vector, ClusterNum):
    a = Vector.shape[1]
    b = []
    NewCluster = np.zeros((1, a))
    NewCluster = np.delete(NewCluster, 0, 0)   
    for i in range(0, ClusterNum, 1):
        b.append(Vector[i])
        NewCluster = np.vstack((NewCluster, b))
        b = []
        
    for i in range(0, ClusterNum, 1):
        Vector = np.delete(Vector, 0, 0)
        
    return (Vector, NewCluster)

def var(n, m):
    aux = []
    for i in range(0, len(n), 1):
        aux.append(sheet.cell(row=n[i], column=m).value)
    return aux

def Procesamiento (Min, Max, Start, Stop, Var):
    """
    Funcion
    """
    x = Var
    
    Tabla, N = tabla(Min, Max, Start, Stop, x)
    
    min_max_scaler = preprocessing.MinMaxScaler()
    X_train_minmax = min_max_scaler.fit_transform(Tabla)*10
    Tabla = preprocessing.normalize(X_train_minmax, norm='l2')
    
    """
    Clustering
    """
    ms = MeanShift()
    n_components = 2
    n_neighbors = 10
    ms.fit(Tabla)
    labels = ms.labels_
    cluster_centers = ms.cluster_centers_
    n_clusters_ = len(np.unique(labels))
    print("Number of estimated clusters:", n_clusters_)
    
    """
    Guardar los Clusters
    """
    f = open('file.txt', 'wb')
    for i in range(len(labels)):
        f.write("%i \n" % labels[i])
    f.close()
    
    """
    TSNE
    """
    Combinadas = armar(Tabla, cluster_centers)
    tsne = manifold.TSNE(n_components=n_components, init='pca', random_state=0)
    Desdoblada = tsne.fit_transform(Combinadas)
    
    
    g, h = desarmar(Desdoblada, n_clusters_)
    
    Var = var(N, x)
    
    plt.scatter(g[:, 0], g[:, 1], c=Var)
        
    for i in range(len(h)):
        plt.plot(h[i][0], h[i][1], c='r', marker='x', markersize = 10) 
    plt.axis('tight')
    plt.show()