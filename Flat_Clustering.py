import numpy as np
import matplotlib.pyplot as plt
from matplotlib import style
style.use("ggplot")
from sklearn.cluster import KMeans
import openpyxl


wb = openpyxl.load_workbook('DataBaseBolivia.xlsx')
sheet = wb.get_sheet_by_name('Pooled data Bolivia')

x = []
y = []
a = []
X = np.array([[0,0],[0,0]])
X=np.delete(X, 0, 0)
X=np.delete(X, 0, 0)

for i in range(2, 1875, 1):
	if sheet.cell(row=i, column=76).value is not None and sheet.cell(row=i, column=417).value is not None:
		x.append(sheet.cell(row=i, column=76).value)
		y.append(sheet.cell(row=i, column=417).value)
		a=[sheet.cell(row=i, column=76).value,sheet.cell(row=i, column=417).value]
		X = np.insert(X,0, a, axis=0)

print 'Cantidad de datos utilizados: ', len(X)

## MOSTRAR LA GRAFICA ANTES APLICAR CLUSTERING
#plt.scatter(x,y)
#plt.show()

kmeans = KMeans(n_clusters=4)
kmeans.fit(X)

centroids = kmeans.cluster_centers_
labels = kmeans.labels_

print(centroids)
print(labels)

colors = ["g.","r.","c.","y."]

for i in range(len(X)):
    #print("coordinate:",X[i], "label:", labels[i])
    plt.plot(X[i][0], X[i][1], colors[labels[i]], markersize = 10)


plt.scatter(centroids[:, 0],centroids[:, 1], marker = "x", s=150, linewidths = 5, zorder = 10)
plt.xlabel("I always avoid risks")
plt.ylabel("planning to buy next cell phone from the same manufacturer")
plt.title("4 Clusters")
plt.show()
